#!/usr/bin/python3
# -*- coding: utf-8 -*- 
# @Author : HEFEI
import os
from textfsm import parser, TextFSM
import time
import threading
import queue
import getpass
from netmiko import ConnectHandler, NetmikoTimeoutException, NetMikoAuthenticationException, ConfigInvalidException, \
    NetMikoTimeoutException
from openpyxl import Workbook, load_workbook
import pandas as pd
# from netmiko import exceptions
import logging


# 创建CMDB表格,初始化；
def INIT_CMDB(log1):
    # 创建配置收集文件夹
    # log2 = 'logs/' + time.strftime('日志_%Y年%m月%d日%H点%M分', time.localtime())
    # if not os.path.exists(log1):
    #     os.makedirs(log1)
    #     print(f'{log1} 文件夹创建完成')
    #     logger.debug(f'{log1} 文件夹创建完成')
    if os.path.exists(f'{log1}/不可达设备.txt'):
        os.remove(f'{log1}/不可达设备.txt')
        logger.debug(f'{log1}/不可达设备.txt创建完成')
    if os.path.exists(f'{log1}/认证失败设备.txt'):
        os.remove(f'{log1}/认证失败设备.txt')
        logger.debug(f'{log1}/认证失败设备.txt创建完成')
    if os.path.exists('CMDB.xlsx'):
        os.remove('CMDB.xlsx')
        logger.debug(f'已删除旧的CMDB.xlsx文件')
    workbook = Workbook('CMDB.xlsx')
    sheet = workbook.create_sheet('主机（NET）')
    # cells=sheet[sheet.dimensions]
    sheet.append([
        'htname',
        'ip',
        'model',
        'type',
        'sn',
        'os',
        'tag',
        'room',
        'location',
        'site',
        'state',
        'U位',
        '上联接口'
    ])
    workbook.save('CMDB.xlsx')
    logger.debug('创建了一个新的CMDB.xlsx文件')


def h3c_tm(ip, text_output, dev):
    if dev == 'hp_comware':
        tm = 'template/H3C_CMDB.template'
    elif dev == 'ruijie_os':
        tm = 'template/ruijie2.tm'
    elif dev == 'huawei':
        tm = 'template/huawei.tm'
    with open(tm) as template:
        fsm = TextFSM(template)
        result = fsm.ParseTextToDicts(text_output)
    # print(fsm.header)
    if len(result) == 0:
        raise ValueError
    data = result[0]
    # 转换上联
    UP_PORT = ''
    UP = zip(data['LOCALPORT'], data['UPDEVICE'], data['UPPORT'])
    if 'DI' in data['HOSTNAME']:
        for j, q, k in UP:
            if 'CO' in q or 'SP' in q or 'EX' in q or 'FW' in q:
                UP_P = f'local-{j} TO {q}_{k}\n'
                UP_PORT += UP_P
    elif 'LF' in data['HOSTNAME']:
        for j, q, k in UP:
            if 'DI' in q or 'CO' in q or 'SP' in q or 'EX' in q or 'FW' in q:
                UP_P = f'local-{j} TO {q}_{k}\n'
                UP_PORT += UP_P
    elif 'CO' in data['HOSTNAME']:
        for j, q, k in UP:
            if 'SP' in q or 'EX' in q or 'EXFW' in q:
                UP_P = f'local-{j} TO {q}_{k}\n'
                UP_PORT += UP_P
    elif 'EX' in data['HOSTNAME']:
        UP_PORT = ''
    elif 'SP' in data['HOSTNAME']:
        for j, q, k in UP:
            if 'CN-PEK51-Z-BI10-IEXNS-00' in q:
                UP_P = f'local-{j} TO {q}_{k}\n'
                UP_PORT += UP_P
    else:
        for j, q, k in UP:
            UP_P = f'local-{j} TO {q}_{k}\n'
            UP_PORT += UP_P
    # 转换SN
    SN = ''
    for s in data['SN']:
        S1 = f'{s},'
        SN += S1
    SN = SN.rstrip(',')
    # 优化LF、DI、CO、SP
    sheet.append([data['HOSTNAME'],
                  ip,
                  data['MODEL'],
                  'sw',
                  SN,
                  data['VERSION'],
                  'cop.xiaomi_owt.it',
                  '',
                  '',
                  '',
                  'working',
                  '',
                  UP_PORT
                  ])
    workbook.save('CMDB.xlsx')
    logger.debug(f"{ip} {data['HOSTNAME']}录入完成")


class ssh_method(threading.Thread):
    def __init__(self, devicety, ip, username, passwd, enablePW, queue):
        threading.Thread.__init__(self)
        # self.Opr = Opr
        self.devicety = devicety
        self.ip = ip
        self.username = username
        self.passwd = passwd
        self.queue = queue
        self.enablePasswd = enablePW

    def run(self):
        try:
            # 生成本地时间
            # date_time = datetime.datetime.now().strftime("%Y-%m-%d")
            # 需要enable密码的厂商
            enableMa = ['ruijie_os', 'cisco_ios']
            h3c_cmdlist = [
                '\n',
                'disp lldp nei list',
                'display version',
                'display device ma']
            ruijie_cmdlist = [
                '\n',
                'show version',
                'show lldp neighbors '
            ]
            huawei_cmdlist = [
                '\n',
                'display lldp neighbor brief',
                'display version',
                'display device ma']
            if self.devicety == 'ruijie_os':
                cmdlist = ruijie_cmdlist
            elif self.devicety == 'hp_comware':
                cmdlist = h3c_cmdlist
            elif self.devicety == 'huawei':
                cmdlist = huawei_cmdlist
            LoginInfo = {'device_type': self.devicety,
                         'host': self.ip,
                         'username': self.username,
                         'password': self.passwd,
                         # 'secret': self.enablePasswd,
                         'port': 22}
            if self.devicety in enableMa and pd.isnull(self.enablePasswd) is False:
                LoginInfo.update({'secret': str(self.enablePasswd).strip()})
            elif self.devicety not in enableMa and pd.isnull(self.enablePasswd) is False:
                raise ValueError(f'{self.devicety} 设备厂商无enable选项。')
            with ConnectHandler(**LoginInfo) as connect:
            # with ConnectHandler(device_type=self.devicety, ip=self.ip, username=self.username,
            #                   password=self.passwd) as connect:
                if self.devicety in enableMa and pd.isnull(self.enablePasswd) is False:
                    connect.enable()
                logger.debug("已经成功登陆交换机" + self.ip)
                time.sleep(2)
                # if os.path.exists(log1 + "/" + self.ip + "-" + date_time + ".txt"):
                #     os.remove(log1 + "/" + self.ip + "-" + date_time + ".txt")
                # if os.path.exists(error_log + "/" + self.ip + "-" + date_time + ".txt"):
                #     os.remove(error_log + "/" + self.ip + "-" + date_time + ".txt")
                # show命令
                output = ''
                for cmdline in cmdlist:
                    output += connect.send_command_timing(cmdline,
                                                          strip_prompt=False,
                                                          strip_command=False
                                                          )
                    # output += connect.send_command(cmdline, strip_prompt=False, strip_command=False, read_timeout=60)
                    # print(output)
                    # 保存文本，后期可以考虑不保存；但是不知道内存会不会有影响；
                with open(f'{log1}/{self.ip}.txt', "a+") as log:
                    log.writelines(output)
                # if self.devicety=='hp_comware':
                h3c_tm(self.ip, output, self.devicety)
                # elif self.devicety=='ruijie':
                #     ruijie_tm(self.ip, output)

        except (NetmikoTimeoutException, EOFError):
            # print('\033[1;31m ***** %s 无法连接设备! \033[0m' % self.ip + '\n')
            logger.error('%s 无法连接设备!' % self.ip + '\n')
            with open(f'{log1}/登陆超时设备.txt', 'a+', encoding='utf-8') as timeout:
                timeout.write(self.ip + ' 连接设备超时!\n')
                logger.error(self.ip + ' 连接设备超时!')
        except NetMikoAuthenticationException:
            # print('\033[1;31m *****%s 用户名或密码错误! \033[0m' % self.ip + '\n')
            logger.error('%s 用户名或密码错误!' % self.ip)
            with open(f'{log1}/认证失败设备.txt', 'a+', encoding='utf-8') as unauth:
                unauth.write(self.ip + ' 用户名密码错误\n')
                # logger.error(self.ip + ' 用户名密码错误')
        except ConfigInvalidException as conf_e:
            with open(f'{log1}/配置失败设备.txt', 'a+', encoding='utf-8') as unconfig:
                unconfig.write(self.ip + ' 型号不匹配导致配置错误\n' + conf_e + '\n')
                # logger.debug(self.ip + ' 型号不匹配导致配置错误\n' + conf_e)
            # print('\033[1;31m ***** {} 型号不匹配导致配置错误\n \033[0m'.format(self.ip), conf_e)
            logger.error('{} 型号不匹配导致配置错误'.format(self.ip), conf_e)
        # except exceptions.ReadTimeout:
        #     print(f'{self.ip}读回显超时')
        except ValueError as v:
            # print(f'设备 {self.ip} 模板 {self.devicety} textfsm模板错误，请检查模板。')
            # if 'Router prompt not found' in v:
            #     logger.error(f'设备 {self.ip} 设备厂商不匹配！', exc_info=True)
            logger.error(f'设备 {self.ip} {str(ValueError)}', exc_info=True)
            # logger.debug(f'设备 {self.ip} 模板 {self.devicety} textfsm模板错误，请检查模板。也有可能账户被锁定', exc_info=True)
            # logger.error(f'设备 {self.ip} 模板 {self.devicety} textfsm模板错误，请检查模板。')
            # with open(f'{log1}/匹配失败设备.txt', 'a+', encoding='utf-8') as muban:
            #     muban.write(f'设备 {self.ip} 模板 {self.devicety} textfsm模板错误，请检查模板。\n')
            #     logger.error(f'设备 {self.ip} 模板 {self.devicety} textfsm模板错误，请检查模板。\n')
            #     logger.debug(f'设备 {self.ip} 模板 {self.devicety} textfsm模板错误，请检查模板或表格信息是否错误。\n',exc_info=True)
        except Exception:
            logger.exception(f'设备 {self.ip} 模板 {self.devicety} 请相关开发人员。\n', exc_info=True)
        finally:
            self.queue.get()
            self.queue.task_done()
            # logger.debug(f'{self.ip}执行完毕!')


# 使用ip表数据登陆操作
def SSH_BY_DB(ip_list):
    # 设置多线程限制数量
    que = queue.Queue(60)
    th_list = []
    for line in ip_list:
        devicety = line['device_type']
        ip = str(line['ip_address']).strip()
        enablePW = line['enable']
        que.put(line)
        t = ssh_method(devicety, ip, username, passwd, enablePW, que)
        t.start()
        th_list.append(t)
    for th in th_list:
        th.join()


if __name__ == '__main__':
    begin_time = time.time()
    # 初始化报错日志文件
    log1 = 'logs/' + time.strftime('日志_%Y年%m月%d日%H点%M分', time.localtime())
    if not os.path.exists(log1):
        os.makedirs(log1)
        print(f'{log1} 文件夹创建完成')
        logging.debug(f'{log1} 文件夹创建完成')

    # logger
    logger = logging.Logger('logger')
    logger.setLevel(logging.DEBUG)
    # log处理器
    consoleHandler = logging.StreamHandler()
    consoleHandler.setLevel(logging.ERROR)
    fileHandler = logging.FileHandler(filename=f'{log1}/logs.log', mode='a+', encoding='utf-8')
    # log formate
    formatters = logging.Formatter(fmt="%(asctime)s %(thread)d %(threadName)s %(levelname)s %(message)s",
                                   datefmt="%Y-%m-%d %H:%M-%S")
    consoleHandler.setFormatter(formatters)
    fileHandler.setFormatter(formatters)
    # logger绑定处理器
    logger.addHandler(consoleHandler)
    logger.addHandler(fileHandler)
    # 初始化CMDB文件和报错日志文件
    INIT_CMDB(log1)
    # 读取ip列表和命令列表
    df = pd.read_excel('ip_info.xlsx', sheet_name='ip')
    # df2 = pd.read_excel('ip_info.xlsx', sheet_name='cmd')
    ip_list = df.to_dict('records')
    # cmdlist = df2['command']
    # cmds = []
    # 打开ip文件，获取ip
    #
    # with open('ip.txt', 'r') as ips:
    #     ip_list = ips.readlines()
    # print('-' * 50)
    print('\n' * 1)
    print('-' * 50)
    print('\033[1;31m请检查ip是否正确！\033[0m')
    host_count = 0
    for i in ip_list:
        host_count += 1
        # print(i)
    print(df)
    logger.debug("操作项目:\n" + str(df))
    print('\n' * 2, '共计 \033[1;31m{}\033[0m 台'.format(host_count))
    logger.debug('共计{}台'.format(host_count))
    print('-' * 50)
    username = input('请输入用户名:')
    logger.debug(f'操作用户:{username}')
    passwd = getpass.getpass('请输入密码:')
    # 打开excel表
    workbook = load_workbook('CMDB.xlsx')
    sheet = workbook['主机（NET）']
    # SSH并通过textfsm抓取关键字写入excel
    SSH_BY_DB(ip_list)
    end_time = time.time()
    cost_time = round(end_time - begin_time)
    print('\n' * 2)
    print('*' * 50)
    if cost_time < 60:
        print('总共花费：{} 秒'.format(cost_time))
        logger.debug('总共花费：{} 秒'.format(cost_time))
    else:
        mini = round(cost_time / 60)
        sec = cost_time % 60
        print('总共花费：{0} 分 {1} 秒'.format(mini, sec))
        logger.debug('总共花费：{0} 分 {1} 秒'.format(mini, sec))
    print('*' * 50)
